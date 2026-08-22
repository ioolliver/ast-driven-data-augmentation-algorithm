import importlib.util
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


def load_comparison_module():
    module_path = (
        Path(__file__).resolve().parents[1]
        / "data"
        / "geo_dataset"
        / "compare_augmentation_methods.py"
    )
    spec = importlib.util.spec_from_file_location(
        "compare_geo_augmentation_methods", module_path
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def build_row(row_id, level="Facil"):
    return {
        "id": row_id,
        "question": f"Pergunta {row_id}",
        "sql_code": f"SELECT {row_id}",
        "level": level,
        "source": "base_dataset",
    }


class CompareGeoAugmentationMethodsTest(unittest.TestCase):
    def test_run_comparison_writes_separate_json_and_workbook_for_each_method(self):
        module = load_comparison_module()
        dataset_rows = [build_row(10), build_row(11, level="Medio")]
        method_augmenters = {
            method.key: self._build_fake_augmenter(method.key)
            for method in module.METHODS
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_dir = Path(tmp_dir) / "outputs"
            input_path = Path(tmp_dir) / "input.json"
            input_path.write_text(
                json.dumps(dataset_rows, ensure_ascii=False), encoding="utf-8"
            )

            outputs = module.run_comparison(
                dataset_path=input_path,
                output_dir=output_dir,
                method_augmenters=method_augmenters,
                max_workers=2,
            )

            self.assertEqual(set(outputs), {method.key for method in module.METHODS})
            self.assertEqual(len(list(output_dir.glob("*.json"))), 3)
            self.assertEqual(len(list(output_dir.glob("*.xlsx"))), 3)

            for method in module.METHODS:
                json_path, workbook_path = outputs[method.key]
                self.assertEqual(
                    json_path.name, f"geo_dataset_{method.key}_augmented.json"
                )
                self.assertEqual(
                    workbook_path.name, f"geo_dataset_{method.key}_augmented.xlsx"
                )

                rows = json.loads(json_path.read_text(encoding="utf-8"))
                self.assertEqual([row["id"] for row in rows], [10, 11])
                self.assertEqual(rows[0]["level"], "Facil")
                self.assertEqual(
                    rows[0]["changed_question"],
                    f"Pergunta 10 [{method.key}]",
                )

                worksheet = load_workbook(workbook_path)["Augmented Pairs"]
                self.assertEqual(
                    [cell.value for cell in worksheet[1]],
                    [
                        "ID",
                        "Level",
                        "Original Question",
                        "Changed Question",
                        "Original SQL",
                        "Changed SQL",
                    ],
                )
                self.assertEqual(worksheet.cell(2, 1).value, 10)
                self.assertEqual(
                    worksheet.cell(2, 4).value,
                    f"Pergunta 10 [{method.key}]",
                )

    def test_can_skip_paraphrase_only_and_preserve_its_existing_output(self):
        module = load_comparison_module()
        dataset_rows = [build_row(10)]
        remaining_methods = [
            method for method in module.METHODS if method.key != "paraphrase_only"
        ]
        method_augmenters = {
            method.key: self._build_fake_augmenter(method.key)
            for method in remaining_methods
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_dir = Path(tmp_dir) / "outputs"
            output_dir.mkdir()
            input_path = Path(tmp_dir) / "input.json"
            input_path.write_text(
                json.dumps(dataset_rows, ensure_ascii=False), encoding="utf-8"
            )
            existing_paraphrase_path = (
                output_dir / "geo_dataset_paraphrase_only_augmented.json"
            )
            existing_paraphrase_path.write_text(
                '{"completed": true}\n', encoding="utf-8"
            )

            outputs = module.run_comparison(
                dataset_path=input_path,
                output_dir=output_dir,
                method_augmenters=method_augmenters,
                max_workers=1,
                skip_paraphrase_only=True,
            )

            self.assertEqual(
                set(outputs), {method.key for method in remaining_methods}
            )
            self.assertEqual(
                existing_paraphrase_path.read_text(encoding="utf-8"),
                '{"completed": true}\n',
            )
            self.assertFalse(
                (output_dir / "geo_dataset_paraphrase_only_augmented.xlsx").exists()
            )

    def test_paraphrase_only_can_preserve_sql_in_export(self):
        module = load_comparison_module()
        rows = [
            {
                "id": 1,
                "original_question": "Pergunta original",
                "changed_question": "Pergunta parafraseada",
                "original_sql": "SELECT 1",
                "changed_sql": "SELECT 1",
                "level": "Facil",
            }
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "paraphrase.xlsx"
            module.write_augmented_workbook(workbook_path, rows)
            worksheet = load_workbook(workbook_path)["Augmented Pairs"]

        self.assertEqual(worksheet.cell(2, 5).value, "SELECT 1")
        self.assertEqual(worksheet.cell(2, 6).value, "SELECT 1")

    @staticmethod
    def _build_fake_augmenter(method_key):
        def augment(question, sql):
            return f"{question} [{method_key}]", f"{sql} -- {method_key}"

        return augment


if __name__ == "__main__":
    unittest.main()
