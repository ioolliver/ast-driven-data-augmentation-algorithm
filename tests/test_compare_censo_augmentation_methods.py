import importlib.util
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


def load_comparison_module():
    module_path = (
        Path(__file__).resolve().parents[1]
        / "data"
        / "censo_escolar_dataset"
        / "compare_augmentation_methods.py"
    )
    spec = importlib.util.spec_from_file_location(
        "compare_censo_augmentation_methods", module_path
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def build_query(query_id, level="Fácil"):
    return {
        "id": query_id,
        "pergunta_nl": f"Pergunta {query_id}",
        "sql": f"SELECT {query_id}",
        "complexidade": {"nivel": level},
    }


class CompareCensoAugmentationMethodsTest(unittest.TestCase):
    def test_run_comparison_writes_separate_json_and_workbook_for_each_method(self):
        module = load_comparison_module()
        queries = [build_query(10), build_query(11, level="Média")]
        method_augmenters = {
            method.key: self._build_fake_augmenter(method.key)
            for method in module.METHODS
        }

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_dir = Path(tmp_dir) / "outputs"
            input_path = Path(tmp_dir) / "input.json"
            input_path.write_text(
                json.dumps({"queries": queries}, ensure_ascii=False),
                encoding="utf-8",
            )

            outputs = module.run_comparison(
                dataset_path=input_path,
                output_dir=output_dir,
                method_augmenters=method_augmenters,
                max_workers=2,
            )

            self.assertEqual(set(outputs), {method.key for method in module.METHODS})
            self.assertEqual(len(list(output_dir.glob("*.json"))), 3)
            self.assertEqual(len(list(output_dir.glob("*.xlsx"))), 3)

            for method in module.METHODS:
                json_path, workbook_path = outputs[method.key]
                self.assertEqual(
                    json_path.name,
                    f"censo_escolar_{method.key}_augmented.json",
                )
                self.assertEqual(
                    workbook_path.name,
                    f"censo_escolar_{method.key}_augmented.xlsx",
                )

                rows = json.loads(json_path.read_text(encoding="utf-8"))
                self.assertEqual([row["id"] for row in rows], [10, 11])
                self.assertEqual(rows[0]["level"], "Fácil")
                self.assertEqual(
                    rows[0]["changed_question"],
                    f"Pergunta 10 [{method.key}]",
                )

                workbook = load_workbook(workbook_path)
                self.assertEqual(workbook.sheetnames, ["Augmented Pairs"])
                worksheet = workbook["Augmented Pairs"]
                self.assertEqual(
                    [cell.value for cell in worksheet[1]],
                    [
                        "ID",
                        "Level",
                        "Original Question",
                        "Changed Question",
                        "Original SQL",
                        "Changed SQL",
                    ],
                )
                self.assertEqual(worksheet.cell(2, 1).value, 10)
                self.assertEqual(
                    worksheet.cell(2, 4).value,
                    f"Pergunta 10 [{method.key}]",
                )

    def test_paraphrase_only_fake_can_preserve_sql_in_export(self):
        module = load_comparison_module()
        rows = [
            {
                "id": 1,
                "original_question": "Pergunta original",
                "changed_question": "Pergunta parafraseada",
                "original_sql": "SELECT 1",
                "changed_sql": "SELECT 1",
                "level": "Fácil",
            }
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "paraphrase.xlsx"
            module.write_augmented_workbook(workbook_path, rows)

            worksheet = load_workbook(workbook_path)["Augmented Pairs"]

        self.assertEqual(worksheet.cell(2, 5).value, "SELECT 1")
        self.assertEqual(worksheet.cell(2, 6).value, "SELECT 1")

    @staticmethod
    def _build_fake_augmenter(method_key):
        def augment(question, sql):
            return f"{question} [{method_key}]", f"{sql} -- {method_key}"

        return augment


if __name__ == "__main__":
    unittest.main()
