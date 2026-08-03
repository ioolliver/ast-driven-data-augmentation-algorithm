import importlib.util
import tempfile
import unittest
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook


SPREADSHEET_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def load_workbook_module():
    module_path = (
        Path(__file__).resolve().parents[1] / "data" / "analysis_workbook.py"
    )
    spec = importlib.util.spec_from_file_location("analysis_workbook", module_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def read_sheet_names(workbook_path):
    with ZipFile(workbook_path) as workbook:
        workbook_xml = ElementTree.fromstring(workbook.read("xl/workbook.xml"))

    namespace = {"x": SPREADSHEET_NAMESPACE}
    return [
        sheet.attrib["name"]
        for sheet in workbook_xml.findall("x:sheets/x:sheet", namespace)
    ]


def distribution(count=2, average=0.25):
    return {
        "count": count,
        "min": 0.0,
        "max": 0.5,
        "average": average,
        "median": average,
        "standard_deviation": 0.1,
        "p25": 0.1,
        "p75": 0.4,
        "p90": 0.45,
        "p95": 0.475,
    }


class AnalysisWorkbookTest(unittest.TestCase):
    def test_writes_semantic_workbook_with_the_example_sheet_contract(self):
        module = load_workbook_module()
        metadata = {
            "generated_at": "2026-08-03T12:00:00+00:00",
            "input_path": "/tmp/input.json",
            "dataset_label": "Censo Escolar Dataset",
            "row_count": 2,
            "model_id": "jinaai/jina-embeddings-v3",
            "embedding_task": "text-matching",
            "score_formula": "clip(1 - cosine_similarity(original, changed), 0, 1)",
            "model_license": "CC BY-NC 4.0",
        }
        summary = {
            "overall": {
                "sql": distribution(),
                "question": distribution(),
                "combined": distribution(),
            },
            "by_level": {
                "Fácil": {
                    "sql": distribution(),
                    "question": distribution(),
                    "combined": distribution(),
                }
            },
            "unchanged_text": {"sql": 1, "question": 1},
            "bands": {
                comparison: {
                    "[0.0, 0.1)": 1,
                    "[0.9, 1.0]": 1,
                }
                for comparison in ("sql", "question", "combined")
            },
        }
        rows = [
            {
                "row_index": 0,
                "level": "Fácil",
                "sql_variation_score": 0.1,
                "question_variation_score": 0.2,
                "combined_variation_score": 0.15,
            },
            {
                "row_index": 1,
                "level": "Fácil",
                "sql_variation_score": 0.4,
                "question_variation_score": 0.5,
                "combined_variation_score": 0.45,
            },
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "semantic.xlsx"
            module.write_semantic_variation_workbook(
                output_path, metadata, summary, rows
            )

            self.assertEqual(
                read_sheet_names(output_path),
                ["Resumo", "Por Nível", "Distribuição", "Extremos", "Metadados"],
            )
            workbook = load_workbook(output_path, data_only=True)
            self.assertEqual(
                workbook["Resumo"]["A1"].value,
                "Análise de Variação Semântica — Censo Escolar",
            )
            self.assertEqual(workbook["Resumo"]["A6"].value, 2)
            self.assertEqual(workbook["Resumo"]["B6"].value, 0.25)
            self.assertEqual(workbook["Distribuição"]["D7"].value, 1)
            self.assertEqual(len(workbook["Distribuição"]._charts), 1)

    def test_writes_component_workbook_with_the_example_sheet_contract(self):
        module = load_workbook_module()
        metadata = {
            "generated_at": "2026-08-03T12:00:00+00:00",
            "input_path": "/tmp/input.json",
            "dataset_label": "Censo Escolar Dataset",
            "row_count": 1,
            "sql_dialect": "bigquery",
            "score_formula": "changed_component_count / component_total",
        }
        summary = {
            "overall": distribution(count=1),
            "by_level": {"Média": distribution(count=1)},
            "bands": {"[0.0, 0.1)": 0, "[0.9, 1.0]": 1},
            "unchanged_sql": 0,
            "changed_component_families": {"predicate": 2},
        }
        rows = [
            {
                "row_index": 0,
                "level": "Média",
                "component_total": 4,
                "changed_component_count": 2,
                "component_matching_score": 0.5,
            }
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "components.xlsx"
            module.write_component_matching_workbook(
                output_path, metadata, summary, rows
            )

            self.assertEqual(
                read_sheet_names(output_path),
                [
                    "Resumo",
                    "Por Nível",
                    "Distribuição",
                    "Componentes",
                    "Extremos",
                    "Metadados",
                ],
            )
            workbook = load_workbook(output_path, data_only=True)
            self.assertEqual(
                workbook["Resumo"]["A1"].value,
                "Análise de Component Matching — Censo Escolar",
            )
            self.assertEqual(workbook["Resumo"]["A6"].value, 1)
            self.assertEqual(workbook["Componentes"]["A6"].value, "predicate")
            self.assertEqual(workbook["Componentes"]["B6"].value, 2)
            self.assertEqual(len(workbook["Componentes"]._charts), 1)


if __name__ == "__main__":
    unittest.main()
