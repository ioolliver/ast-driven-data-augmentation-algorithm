from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


TITLE_FILL = PatternFill("solid", fgColor="17365D")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="2F75B5")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=18)
SECTION_FONT = Font(color="17365D", bold=True, size=12)
HEADER_FONT = Font(color="FFFFFF", bold=True)
SCORE_FORMAT = "0.000000"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"
TABLE_STYLE = "TableStyleMedium2"
STATISTIC_COLUMNS = (
    ("Count", "count"),
    ("Min", "min"),
    ("Max", "max"),
    ("Average", "average"),
    ("Median", "median"),
    ("Std Dev", "standard_deviation"),
    ("P25", "p25"),
    ("P75", "p75"),
    ("P90", "p90"),
    ("P95", "p95"),
)


def _new_sheet(workbook, sheet_name, heading, subtitle, column_count=12):
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=column_count
    )
    title_cell = worksheet.cell(1, 1, heading)
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 28
    worksheet.merge_cells(
        start_row=3, start_column=1, end_row=3, end_column=column_count
    )
    worksheet.cell(3, 1, subtitle).font = Font(color="666666", italic=True)
    worksheet.sheet_view.showGridLines = False
    return worksheet


def _write_section_title(worksheet, row, title, width):
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = worksheet.cell(row, 1, title)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT


def _write_table(worksheet, start_row, headers, rows, table_name):
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(start_row, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_offset, values in enumerate(rows, start=1):
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(start_row + row_offset, column, value)
            if isinstance(value, float):
                cell.number_format = SCORE_FORMAT

    end_row = start_row + len(rows)
    end_column = len(headers)
    if rows:
        table = Table(
            displayName=table_name,
            ref=f"A{start_row}:{get_column_letter(end_column)}{end_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name=TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    return end_row


def _set_column_widths(worksheet, widths):
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _distribution_values(distribution):
    return [distribution[key] for _, key in STATISTIC_COLUMNS]


def _add_bar_chart(worksheet, min_row, max_row, data_column, title, anchor):
    if max_row < min_row:
        return
    chart = BarChart()
    chart.title = title
    chart.style = 10
    chart.height = 7
    chart.width = 12
    chart.y_axis.title = "Quantidade"
    chart.x_axis.title = "Faixa"
    data = Reference(
        worksheet,
        min_col=data_column,
        min_row=min_row - 1,
        max_row=max_row,
    )
    categories = Reference(
        worksheet,
        min_col=1,
        min_row=min_row,
        max_row=max_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend = None
    worksheet.add_chart(chart, anchor)


def _write_metadata_sheet(workbook, metadata, fields, definition, limitation):
    worksheet = _new_sheet(
        workbook,
        "Metadados",
        "Metadados, definição e limitações",
        "Configuração, definição da métrica e limitações",
        column_count=10,
    )
    rows = []
    for label, key in fields:
        value = metadata[key]
        if key == "generated_at":
            value = _parse_datetime(value)
        rows.append((label, value))
    end_row = _write_table(worksheet, 5, ("Campo", "Valor"), rows, "AnalysisMetadata")
    if isinstance(worksheet["B6"].value, datetime):
        worksheet["B6"].number_format = DATETIME_FORMAT

    definition_row = end_row + 3
    _write_section_title(worksheet, definition_row, "Definição", 10)
    worksheet.merge_cells(
        start_row=definition_row + 1,
        start_column=1,
        end_row=definition_row + 3,
        end_column=10,
    )
    worksheet.cell(definition_row + 1, 1, definition).alignment = Alignment(
        wrap_text=True, vertical="top"
    )

    limitation_row = definition_row + 5
    _write_section_title(worksheet, limitation_row, "Limitação", 10)
    worksheet.merge_cells(
        start_row=limitation_row + 1,
        start_column=1,
        end_row=limitation_row + 4,
        end_column=10,
    )
    worksheet.cell(limitation_row + 1, 1, limitation).alignment = Alignment(
        wrap_text=True, vertical="top"
    )
    _set_column_widths(worksheet, [22, 95])
    worksheet.freeze_panes = "A5"


def _parse_datetime(value):
    if not isinstance(value, str):
        return value
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return value


def _write_extreme_rows(worksheet, headers, lowest_rows, highest_rows, row_mapper):
    table_rows = []
    for group, rows in (("Menor", lowest_rows), ("Maior", highest_rows)):
        for rank, row in enumerate(rows, start=1):
            table_rows.append((group, rank, *row_mapper(row)))
    _write_table(worksheet, 5, headers, table_rows, "AnalysisExtremes")
    worksheet.freeze_panes = "A6"


def _semantic_overall_rows(summary):
    return [
        (comparison.title(), *_distribution_values(summary["overall"][comparison]))
        for comparison in ("sql", "question", "combined")
    ]


def _short_dataset_label(metadata):
    return metadata["dataset_label"].removesuffix(" Dataset")


def write_semantic_variation_workbook(output_path, metadata, summary, scored_rows):
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_sheet = _new_sheet(
        workbook,
        "Resumo",
        f"Análise de Variação Semântica — {_short_dataset_label(metadata)}",
        "Resultados consolidados da análise de variação semântica",
    )
    metric_headers = (
        "Linhas analisadas",
        "Média combinada",
        "Mediana combinada",
        "Perguntas inalteradas",
    )
    metric_values = (
        metadata["row_count"],
        summary["overall"]["combined"]["average"],
        summary["overall"]["combined"]["median"],
        summary["unchanged_text"]["question"],
    )
    _write_table(summary_sheet, 5, metric_headers, [metric_values], "SemanticHeadline")
    _write_section_title(summary_sheet, 9, "Estatísticas gerais", 11)
    statistic_headers = ("Comparison", *(label for label, _ in STATISTIC_COLUMNS))
    _write_table(
        summary_sheet,
        10,
        statistic_headers,
        _semantic_overall_rows(summary),
        "SemanticOverallStats",
    )
    _write_section_title(summary_sheet, 15, "Score combinado por nível", 6)
    level_rows = [
        (
            level,
            values["combined"]["count"],
            values["combined"]["average"],
            values["combined"]["median"],
            values["combined"]["p90"],
            values["combined"]["p95"],
        )
        for level, values in summary["by_level"].items()
    ]
    _write_table(
        summary_sheet,
        16,
        ("Nível", "Count", "Average", "Median", "P90", "P95"),
        level_rows,
        "SemanticCombinedByLevel",
    )
    _set_column_widths(summary_sheet, [22, 16, 18, 18, 18, 18, 18, 18, 18, 18, 18])
    summary_sheet.freeze_panes = "A5"

    level_sheet = _new_sheet(
        workbook,
        "Por Nível",
        "Estatísticas por nível",
        "SQL, pergunta e score combinado por dificuldade",
    )
    level_detail_rows = []
    for level, comparisons in summary["by_level"].items():
        for comparison in ("sql", "question", "combined"):
            level_detail_rows.append(
                (
                    level,
                    comparison.title(),
                    *_distribution_values(comparisons[comparison]),
                )
            )
    _write_table(
        level_sheet,
        5,
        ("Nível", "Comparison", *(label for label, _ in STATISTIC_COLUMNS)),
        level_detail_rows,
        "SemanticByLevel",
    )
    _set_column_widths(level_sheet, [22, 16, *([14] * 10)])
    level_sheet.freeze_panes = "A6"

    distribution_sheet = _new_sheet(
        workbook,
        "Distribuição",
        "Distribuição por faixas de score",
        "Contagem para SQL, pergunta e score combinado",
    )
    band_labels = list(summary["bands"]["combined"])
    band_rows = [
        (
            label,
            summary["bands"]["sql"].get(label, 0),
            summary["bands"]["question"].get(label, 0),
            summary["bands"]["combined"].get(label, 0),
        )
        for label in band_labels
    ]
    end_row = _write_table(
        distribution_sheet,
        5,
        ("Score Band", "SQL", "Question", "Combined"),
        band_rows,
        "SemanticScoreBands",
    )
    _add_bar_chart(distribution_sheet, 6, end_row, 4, "Score combinado", "F5")
    _set_column_widths(distribution_sheet, [20, 14, 14, 14])
    distribution_sheet.freeze_panes = "A6"

    extremes_sheet = _new_sheet(
        workbook,
        "Extremos",
        "Menores e maiores scores combinados",
        "Linhas com menor e maior variação semântica agregada",
    )
    sorted_rows = sorted(
        scored_rows,
        key=lambda row: (row["combined_variation_score"], row["row_index"]),
    )
    _write_extreme_rows(
        extremes_sheet,
        (
            "Grupo",
            "Rank",
            "Row Index",
            "Level",
            "SQL Score",
            "Question Score",
            "Combined Score",
        ),
        sorted_rows[:10],
        list(reversed(sorted_rows[-10:])),
        lambda row: (
            row["row_index"],
            row["level"],
            row["sql_variation_score"],
            row["question_variation_score"],
            row["combined_variation_score"],
        ),
    )
    _set_column_widths(extremes_sheet, [14, 10, 14, 20, 16, 20, 20])

    _write_metadata_sheet(
        workbook,
        metadata,
        (
            ("Generated at", "generated_at"),
            ("Input", "input_path"),
            ("Rows analyzed", "row_count"),
            ("Embedding model", "model_id"),
            ("Embedding task", "embedding_task"),
            ("Score formula", "score_formula"),
            ("Model license", "model_license"),
        ),
        "A score of 0 represents no detected semantic variation in embedding space; "
        "a score of 1 represents maximum variation under the clipped cosine metric.",
        "This score is an embedding-based heuristic for variation strength. It does "
        "not prove SQL behavioral equivalence or difference.",
    )
    workbook.save(output_path)


def write_component_matching_workbook(output_path, metadata, summary, scored_rows):
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary_sheet = _new_sheet(
        workbook,
        "Resumo",
        f"Análise de Component Matching — {_short_dataset_label(metadata)}",
        "Resultados consolidados da análise estrutural dos componentes SQL",
    )
    _write_table(
        summary_sheet,
        5,
        ("Linhas analisadas", "Score médio", "Mediana", "SQL sem mudanças"),
        [
            (
                metadata["row_count"],
                summary["overall"]["average"],
                summary["overall"]["median"],
                summary["unchanged_sql"],
            )
        ],
        "ComponentHeadline",
    )
    _write_section_title(summary_sheet, 9, "Estatísticas gerais", 10)
    _write_table(
        summary_sheet,
        10,
        tuple(label for label, _ in STATISTIC_COLUMNS),
        [_distribution_values(summary["overall"])],
        "ComponentOverallStats",
    )
    _write_section_title(summary_sheet, 14, "Resumo por nível", 6)
    overview_rows = [
        (
            level,
            values["count"],
            values["average"],
            values["median"],
            values["p90"],
            values["p95"],
        )
        for level, values in summary["by_level"].items()
    ]
    _write_table(
        summary_sheet,
        15,
        ("Nível", "Count", "Average", "Median", "P90", "P95"),
        overview_rows,
        "ComponentCombinedByLevel",
    )
    _set_column_widths(summary_sheet, [22, *([16] * 9)])
    summary_sheet.freeze_panes = "A5"

    level_sheet = _new_sheet(
        workbook,
        "Por Nível",
        "Estatísticas por nível",
        "Comparação das distribuições de component_matching_score",
    )
    _write_table(
        level_sheet,
        5,
        ("Nível", *(label for label, _ in STATISTIC_COLUMNS)),
        [
            (level, *_distribution_values(values))
            for level, values in summary["by_level"].items()
        ],
        "ComponentByLevel",
    )
    _set_column_widths(level_sheet, [22, *([14] * 10)])
    level_sheet.freeze_panes = "A6"

    distribution_sheet = _new_sheet(
        workbook,
        "Distribuição",
        "Distribuição por faixas de score",
        "Contagem de linhas em cada intervalo",
    )
    band_rows = list(summary["bands"].items())
    end_row = _write_table(
        distribution_sheet,
        5,
        ("Score Band", "Rows"),
        band_rows,
        "ComponentScoreBands",
    )
    _add_bar_chart(distribution_sheet, 6, end_row, 2, "Distribuição dos scores", "D5")
    _set_column_widths(distribution_sheet, [20, 14])
    distribution_sheet.freeze_panes = "A6"

    component_sheet = _new_sheet(
        workbook,
        "Componentes",
        "Famílias de componentes alteradas",
        "Frequência de mudança por família do AST normalizado",
    )
    component_rows = list(summary["changed_component_families"].items()) or [
        ("none", 0)
    ]
    component_end_row = _write_table(
        component_sheet,
        5,
        ("Component Family", "Changed Count"),
        component_rows,
        "ChangedComponentFamilies",
    )
    _add_bar_chart(
        component_sheet,
        6,
        component_end_row,
        2,
        "Componentes alterados",
        "D5",
    )
    _set_column_widths(component_sheet, [24, 18])
    component_sheet.freeze_panes = "A6"

    extremes_sheet = _new_sheet(
        workbook,
        "Extremos",
        "Menores e maiores scores",
        "Linhas com menor e maior variação estrutural",
    )
    sorted_rows = sorted(
        scored_rows,
        key=lambda row: (row["component_matching_score"], row["row_index"]),
    )
    _write_extreme_rows(
        extremes_sheet,
        ("Grupo", "Rank", "Row Index", "Level", "Components", "Changed", "Score"),
        sorted_rows[:10],
        list(reversed(sorted_rows[-10:])),
        lambda row: (
            row["row_index"],
            row["level"],
            row["component_total"],
            row["changed_component_count"],
            row["component_matching_score"],
        ),
    )
    _set_column_widths(extremes_sheet, [14, 10, 14, 20, 16, 14, 16])

    _write_metadata_sheet(
        workbook,
        metadata,
        (
            ("Generated at", "generated_at"),
            ("Input", "input_path"),
            ("Rows analyzed", "row_count"),
            ("SQL dialect", "sql_dialect"),
            ("Score formula", "score_formula"),
        ),
        "component_matching_score is the share of normalized SQL AST component slots "
        "that changed between the original and augmented SQL.",
        "This score is a structural-change heuristic. It does not prove SQL "
        "correctness, behavioral equivalence, or natural-language alignment.",
    )
    workbook.save(output_path)
