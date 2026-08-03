import argparse
import importlib.util
import logging
import sys
from pathlib import Path
from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


LOGGER = logging.getLogger(__name__)
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[1]
INPUT_DATASET_PATH = SCRIPT_DIR / "original_dataset.json"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "augmentation_method_results"
DEFAULT_MAX_WORKERS = 5
BATCH_MODULE_PATH = SCRIPT_DIR / "apply_augmentation_censo_escolar_dataset.py"
SCHEMA_MODULE_PATH = SCRIPT_DIR / "schema.py"


class MethodConfig(NamedTuple):
    key: str
    label: str


METHODS = (
    MethodConfig("paraphrase_only", "Question paraphrasing only"),
    MethodConfig("algorithm_only", "AST algorithm augmentation only"),
    MethodConfig(
        "algorithm_with_paraphrasing",
        "AST algorithm augmentation with question paraphrasing",
    ),
)

HEADERS = (
    "ID",
    "Level",
    "Original Question",
    "Changed Question",
    "Original SQL",
    "Changed SQL",
)
HEADER_FILL = PatternFill("solid", fgColor="2F75B5")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TABLE_STYLE = "TableStyleMedium2"


def load_batch_module():
    spec = importlib.util.spec_from_file_location(
        "censo_escolar_augmentation_batch", BATCH_MODULE_PATH
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def load_schema():
    spec = importlib.util.spec_from_file_location(
        "censo_escolar_comparison_schema", SCHEMA_MODULE_PATH
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.censo_escolar_schema


def build_method_augmenters(schema):
    if str(REPO_ROOT) not in sys.path:
        sys.path.insert(0, str(REPO_ROOT))

    from augmentor import (
        create_paraphrase_only_variation,
        create_random_variation,
        create_random_variation_with_paraphrasing,
    )

    def algorithm_only(question, sql):
        return create_random_variation(schema, question, sql)

    def algorithm_with_paraphrasing(question, sql):
        return create_random_variation_with_paraphrasing(schema, question, sql)

    return {
        "paraphrase_only": create_paraphrase_only_variation,
        "algorithm_only": algorithm_only,
        "algorithm_with_paraphrasing": algorithm_with_paraphrasing,
    }


def add_query_ids(queries, changed_rows):
    return [
        {
            "id": query["id"],
            "original_question": changed_row["original_question"],
            "original_sql": changed_row["original_sql"],
            "changed_question": changed_row["changed_question"],
            "changed_sql": changed_row["changed_sql"],
            "level": changed_row["level"],
        }
        for query, changed_row in zip(queries, changed_rows, strict=True)
    ]


def write_augmented_workbook(output_path, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Augmented Pairs"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    worksheet.append(HEADERS)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        worksheet.append(
            (
                row["id"],
                row["level"],
                row["original_question"],
                row["changed_question"],
                row["original_sql"],
                row["changed_sql"],
            )
        )

    for worksheet_row in worksheet.iter_rows(min_row=2):
        for cell in worksheet_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if rows:
        table = Table(
            displayName="AugmentedPairsTable",
            ref=f"A1:F{len(rows) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name=TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    column_widths = (12, 14, 55, 55, 75, 75)
    for column, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width

    workbook.save(output_path)


def _output_paths(output_dir, method_key):
    filename = f"censo_escolar_{method_key}_augmented"
    return output_dir / f"{filename}.json", output_dir / f"{filename}.xlsx"


def run_comparison(
    dataset_path=INPUT_DATASET_PATH,
    output_dir=DEFAULT_OUTPUT_DIR,
    schema=None,
    method_augmenters=None,
    max_workers=DEFAULT_MAX_WORKERS,
):
    if max_workers <= 0:
        raise ValueError("max_workers must be greater than zero")

    batch = load_batch_module()
    queries = batch.load_queries(dataset_path)
    if method_augmenters is None:
        active_schema = schema if schema is not None else load_schema()
        active_augmenters = build_method_augmenters(active_schema)
    else:
        active_augmenters = method_augmenters

    missing_methods = [
        method.key for method in METHODS if method.key not in active_augmenters
    ]
    if missing_methods:
        raise ValueError(
            "Missing augmenters for methods: " + ", ".join(missing_methods)
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    outputs = {}

    for method in METHODS:
        augment_pair = active_augmenters[method.key]

        LOGGER.info(
            "Starting method: key=%s label=%s queries=%d max_workers=%d",
            method.key,
            method.label,
            len(queries),
            max_workers,
        )
        try:
            _, changed_rows = batch.build_augmented_outputs(
                queries,
                augment_pair,
                max_workers=max_workers,
                progress_callback=batch.log_progress,
            )
        except Exception as exc:
            raise RuntimeError(
                f"Augmentation method {method.key} failed: {exc}"
            ) from exc

        augmented_rows = add_query_ids(queries, changed_rows)
        json_path, workbook_path = _output_paths(output_dir, method.key)
        batch.write_json(json_path, augmented_rows)
        write_augmented_workbook(workbook_path, augmented_rows)
        outputs[method.key] = (json_path, workbook_path)
        LOGGER.info(
            "Completed method: key=%s augmented_pairs=%d json=%s xlsx=%s",
            method.key,
            len(augmented_rows),
            json_path,
            workbook_path,
        )

    return outputs


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    parser = argparse.ArgumentParser(
        description=(
            "Run the three Censo Escolar augmentation methods and write a separate "
            "JSON and XLSX file for each method."
        )
    )
    parser.add_argument("--input", type=Path, default=INPUT_DATASET_PATH)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--max-workers",
        type=int,
        default=DEFAULT_MAX_WORKERS,
        help=(
            "Maximum simultaneous LLM requests within each method "
            f"(default: {DEFAULT_MAX_WORKERS})."
        ),
    )
    args = parser.parse_args()
    run_comparison(
        dataset_path=args.input,
        output_dir=args.output_dir,
        max_workers=args.max_workers,
    )


if __name__ == "__main__":
    main()
