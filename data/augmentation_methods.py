from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class MethodConfig(NamedTuple):
    key: str
    label: str


METHODS = (
    MethodConfig("paraphrase_only", "Question paraphrasing only"),
    MethodConfig("algorithm_only", "AST algorithm augmentation only"),
    MethodConfig(
        "algorithm_with_paraphrasing",
        "AST algorithm augmentation with question paraphrasing",
    ),
)

HEADERS = (
    "ID",
    "Level",
    "Original Question",
    "Changed Question",
    "Original SQL",
    "Changed SQL",
)
HEADER_FILL = PatternFill("solid", fgColor="2F75B5")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TABLE_STYLE = "TableStyleMedium2"


def write_augmented_workbook(output_path, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Augmented Pairs"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    worksheet.append(HEADERS)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        worksheet.append(
            (
                row["id"],
                row["level"],
                row["original_question"],
                row["changed_question"],
                row["original_sql"],
                row["changed_sql"],
            )
        )

    for worksheet_row in worksheet.iter_rows(min_row=2):
        for cell in worksheet_row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if rows:
        table = Table(
            displayName="AugmentedPairsTable",
            ref=f"A1:F{len(rows) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name=TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    column_widths = (12, 14, 55, 55, 75, 75)
    for column, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width

    workbook.save(output_path)
